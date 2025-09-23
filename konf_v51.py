import streamlit as st
import pandas as pd
from ortools.sat.python import cp_model
from io import BytesIO
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Border, Side
import math
import time
import re

st.set_page_config(layout="wide")
st.title("📊 Konferenzplaner (V3)")

# Session State initialisieren (nur für Plan 1)
if "df_plan_1" not in st.session_state:
    st.session_state.df_plan_1 = None
if "df_warte_1" not in st.session_state:
    st.session_state.df_warte_1 = None
if "lehrer_einsaetze_1" not in st.session_state:
    st.session_state.lehrer_einsaetze_1 = None
if "df_klassensicht_1" not in st.session_state:
    st.session_state.df_klassensicht_1 = None
if "klassenlehrer_map" not in st.session_state:
    st.session_state.klassenlehrer_map = None
if "klassenlehrer_map_full" not in st.session_state:
    st.session_state.klassenlehrer_map_full = None
if "statistik_1" not in st.session_state:
    st.session_state.statistik_1 = None
if "berechnung_abgeschlossen" not in st.session_state:
    st.session_state.berechnung_abgeschlossen = False

# Sidebar für Uploads und Einstellungen
with st.sidebar:
    st.header("📁 Daten hochladen")
    lehrer_file = st.file_uploader("Lehrer-Klassen-Zuordnung (CSV)", type=["csv"])
    klassenlehrer_file = st.file_uploader("Klassenlehrer-Zuordnung (CSV)", type=["csv"])

    with st.expander("⚙️ Einstellungen"):
        jahrgaenge_input = st.text_input("Jahrgänge (z.B. 5-9)", value="5-9")
        startzeit_input = st.text_input("Startzeit (HH:MM)", value="14:20")
        dauer_konf = st.number_input("Dauer Konferenz (min)", value=20, min_value=1)
        dauer_pause = st.number_input("Dauer Pause (min)", value=5, min_value=0)
        anzahl_parallel = st.number_input("Anzahl Parallele", value=3, min_value=1)
        raeume_input = st.text_input("Räume (kommagetrennt)", value="109, 111, 112")

# Copyright-Vermerk
    st.markdown("---")
    st.markdown("<p style='font-size: 12px; color: #666;'>\u00A9 2025 N. Klietsch</p>", unsafe_allow_html=True)

# Funktion zur Erstellung eines Konferenzplans
def erstelle_konferenzplan(df_lehrer_filtered, klassenlehrer_map_full, plan_nummer, startzeit_str, dauer_konf, dauer_pause, anzahl_parallel, raeume_list):
    if df_lehrer_filtered.empty:
        st.error(f"Keine Klassen für Plan {plan_nummer} gefunden.")
        return None, None, None, None, None
    
    # Lehrer pro Klasse (einmalig, Duplikate entfernen)
    df_lehrer_unique = df_lehrer_filtered.drop_duplicates(subset=["Klasse", "Lehrer"])
    
    klassen = sorted(df_lehrer_unique["Klasse"].unique())
    lehrer_map = df_lehrer_unique.groupby("Lehrer")["Klasse"].apply(set).to_dict()
    
    # Klassenlehrer ergänzen (nur wenn sie Klassen haben)
    for klasse, kl_lehrer_liste in klassenlehrer_map_full.items():
        if klasse in klassen:
            for kl_lehrer in kl_lehrer_liste:
                if kl_lehrer not in lehrer_map:
                    lehrer_map[kl_lehrer] = set()
                lehrer_map[kl_lehrer].add(klasse)

    # Lehrer ohne Klassen entfernen
    lehrer_map = {lehrer: klassen for lehrer, klassen in lehrer_map.items() if klassen}

    total_klassen = len(klassen)
    max_slots = 12
    total_slots = min(max_slots, math.ceil(total_klassen / anzahl_parallel))
    zeitslots = list(range(total_slots))

    if total_klassen > max_slots * anzahl_parallel:
        st.error(f"Zu viele Klassen ({total_klassen}) für Plan {plan_nummer}. Maximal möglich: {max_slots * anzahl_parallel}.")
        return None, None, None, None, None

    # Statusbalken initialisieren
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    model = cp_model.CpModel()
    slot_var = {k: model.NewIntVar(0, total_slots - 1, f"slot_{k}") for k in klassen}
    raum_var = {k: model.NewIntVar(0, anzahl_parallel - 1, f"raum_{k}") for k in klassen}

    # Status aktualisieren
    status_text.text(f"Plan {plan_nummer}: Modell wird erstellt...")
    progress_bar.progress(10)

    # Slots
    in_slot_vars = {}
    for s in zeitslots:
        for k in klassen:
            in_slot_vars[(s, k)] = model.NewBoolVar(f"in_slot_{s}_{k}")
            model.Add(slot_var[k] == s).OnlyEnforceIf(in_slot_vars[(s, k)])
            model.Add(slot_var[k] != s).OnlyEnforceIf(in_slot_vars[(s, k)].Not())

    # Status aktualisieren
    status_text.text(f"Plan {plan_nummer}: Raumbelegung wird konfiguriert...")
    progress_bar.progress(30)

    # Raumbelegung
    for s in zeitslots:
        klassen_in_slot = [in_slot_vars[(s, k)] for k in klassen]
        if s < total_slots - 1:
            model.Add(sum(klassen_in_slot) == anzahl_parallel)
        else:
            model.Add(sum(klassen_in_slot) <= anzahl_parallel)
        for i, k1 in enumerate(klassen):
            for j, k2 in enumerate(klassen):
                if i < j:
                    both_in_slot = model.NewBoolVar(f"both_{k1}_{k2}_in_slot_{s}")
                    model.AddBoolAnd([in_slot_vars[(s, k1)], in_slot_vars[(s, k2)]]).OnlyEnforceIf(both_in_slot)
                    model.AddBoolOr([in_slot_vars[(s, k1)].Not(), in_slot_vars[(s, k2)].Not()]).OnlyEnforceIf(both_in_slot.Not())
                    model.Add(raum_var[k1] != raum_var[k2]).OnlyEnforceIf(both_in_slot)

    # Status aktualisieren
    status_text.text(f"Plan {plan_nummer}: Klassenlehrer-Zuordnung wird verarbeitet...")
    progress_bar.progress(50)

    # Klassenlehrer - beide Lehrer müssen verschiedene Slots haben
    for klasse, kl_lehrer_liste in klassenlehrer_map_full.items():
        if klasse in klassen:
            for kl_lehrer in kl_lehrer_liste:
                if kl_lehrer in lehrer_map:
                    for andere in lehrer_map[kl_lehrer]:
                        if andere != klasse:
                            model.Add(slot_var[andere] != slot_var[klasse])

    # Status aktualisieren
    status_text.text(f"Plan {plan_nummer}: Konflikte werden minimiert...")
    progress_bar.progress(70)

    # Neue Constraint: Kein Lehrer hat mehr als zwei Klassen in einem Slot
    for lehrer, kl_liste in lehrer_map.items():
        for s in zeitslots:
            klassen_in_slot = [in_slot_vars[(s, k)] for k in kl_liste]
            model.Add(sum(klassen_in_slot) <= 2)

    # Lehrer nicht doppelt gleichzeitig (Minimiere Konflikte, aber nur paarweise für die verbleibenden)
    konfliktvars = []
    for lehrer, kl_liste in lehrer_map.items():
        kl_liste = list(kl_liste)
        for i in range(len(kl_liste)):
            for j in range(i + 1, len(kl_liste)):
                conflict = model.NewBoolVar(f"konflikt_{lehrer}_{kl_liste[i]}_{kl_liste[j]}")
                model.Add(slot_var[kl_liste[i]] == slot_var[kl_liste[j]]).OnlyEnforceIf(conflict)
                model.Add(slot_var[kl_liste[i]] != slot_var[kl_liste[j]]).OnlyEnforceIf(conflict.Not())
                konfliktvars.append(conflict)

    model.Minimize(sum(konfliktvars))
    
    # Status aktualisieren
    status_text.text(f"Plan {plan_nummer}: Lösung wird berechnet...")
    progress_bar.progress(80)
    
    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 120
    solver.parameters.num_search_workers = 4
    status = solver.Solve(model)

    # Status aktualisieren
    progress_bar.progress(90)

    if status in [cp_model.OPTIMAL, cp_model.FEASIBLE]:
        # Startzeit parsen
        start_h, start_m = map(int, startzeit_str.split(":"))
        start_minutes = start_h * 60 + start_m
        
        # Zeitschema berechnen (Start- und Endzeiten pro Slot)
        zeitschema_start = []
        zeitschema_end = []
        current_time = start_minutes
        for i in range(total_slots):
            start_str = f"{current_time // 60:02d}:{current_time % 60:02d}"
            end_time = current_time + dauer_konf
            end_str = f"{end_time // 60:02d}:{end_time % 60:02d}"
            zeitschema_start.append(start_str)
            zeitschema_end.append(end_str)
            current_time = end_time + dauer_pause
        
        plan, lehrer_einsaetze, klassensicht = [], {l: [] for l in lehrer_map}, {k: [] for k in klassen}

        for k in klassen:
            slot = solver.Value(slot_var[k])
            raum_idx = solver.Value(raum_var[k])
            start_zeit = zeitschema_start[slot]
            end_zeit = zeitschema_end[slot]
            raum = raeume_list[raum_idx]
            teilnehmer = [l for l, ks in lehrer_map.items() if k in ks]
            for l in teilnehmer:
                lehrer_einsaetze[l].append((start_zeit, k, raum))
                klassensicht[k].append(l)
            plan.append((k, start_zeit, end_zeit, raum, ", ".join(sorted(teilnehmer))))

        df_plan = pd.DataFrame(plan, columns=["Klasse", "Startzeit", "Endzeit", "Raum", "Lehrkräfte"])
        df_plan = df_plan.sort_values(by="Startzeit")
        
        # Klassensicht mit eindeutigen Lehrern
        df_klassensicht = pd.DataFrame([
            (k, sorted(set(liste))) for k, liste in klassensicht.items()
        ], columns=["Klasse", "Lehrkräfte"])

        # KORRIGIERTE Wartezeiten-Berechnung
        warte_df = []
        for l, einsaetze in lehrer_einsaetze.items():
            if len(einsaetze) > 0:
                # Slot-Nummern aus Startzeiten ableiten
                slot_numbers = []
                for z, k, r in einsaetze:
                    try:
                        slot_index = zeitschema_start.index(z)
                        slot_numbers.append(slot_index)
                    except ValueError:
                        # Fallback: Slot aus Zeit berechnen
                        z_min = int(z.split(":")[0]) * 60 + int(z.split(":")[1])
                        slot_index = math.floor((z_min - start_minutes) / (dauer_konf + dauer_pause))
                        slot_numbers.append(slot_index)
                
                if slot_numbers:
                    min_slot = min(slot_numbers)
                    max_slot = max(slot_numbers)
                    # Gesamtdauer von erstem Start bis letztem Ende + Pausen
                    dauer = max(0, (max_slot - min_slot + 1) * dauer_konf + (max_slot - min_slot) * dauer_pause)
                    anwesenheit = len(einsaetze) * dauer_konf
                    leerlauf = max(0, dauer - anwesenheit)
                else:
                    dauer = anwesenheit = leerlauf = 0
            else:
                dauer = anwesenheit = leerlauf = 0
            
            warte_df.append((l, len(einsaetze), dauer, anwesenheit, leerlauf))
        
        df_warte = pd.DataFrame(warte_df,
                                columns=["Lehrer", "Einsätze", "Gesamtzeit (min)", "Aktive Zeit (min)", "Wartezeit (min)"])

        # Verbesserte Statistik (auch korrigiert)
        df_stat = pd.DataFrame({
            "Einsätze": [len(e) for e in lehrer_einsaetze.values()],
            "Wartezeit": list(df_warte["Wartezeit (min)"])
        })
        
        # Detaillierte Statistik
        statistik = {
            "Allgemein": {
                "Anzahl Klassen": total_klassen,
                "Anzahl Zeitslots": total_slots,
                "Anzahl Lehrer": len(lehrer_map)
            },
            "Einsätze pro Lehrer": {
                "Minimum": df_stat["Einsätze"].min(),
                "Maximum": df_stat["Einsätze"].max(),
                "Durchschnitt": round(df_stat["Einsätze"].mean(), 1),
                "Gesamt": df_stat["Einsätze"].sum()
            },
            "Wartezeiten (Minuten)": {
                "Minimum": max(0, df_stat["Wartezeit"].min()),  # Minimum 0
                "Maximum": df_stat["Wartezeit"].max(),
                "Durchschnitt": round(df_stat["Wartezeit"].mean(), 1),
                "Gesamt": df_stat["Wartezeit"].sum()
            },
            "Verteilung Wartezeiten": {
                "0-20 min": len(df_stat[df_stat["Wartezeit"] <= 20]),
                "21-40 min": len(df_stat[(df_stat["Wartezeit"] > 20) & (df_stat["Wartezeit"] <= 40)]),
                "41-60 min": len(df_stat[(df_stat["Wartezeit"] > 40) & (df_stat["Wartezeit"] <= 60)]),
                "Über 60 min": len(df_stat[df_stat["Wartezeit"] > 60])
            }
        }
        
        # Status aktualisieren
        status_text.text(f"Plan {plan_nummer}: Berechnung abgeschlossen!")
        progress_bar.progress(100)
        time.sleep(0.5)
        progress_bar.empty()
        status_text.empty()
        
        return df_plan, df_warte, lehrer_einsaetze, df_klassensicht, statistik
    else:
        st.error(f"Keine Lösung für Plan {plan_nummer} gefunden.")
        progress_bar.empty()
        status_text.empty()
        return None, None, None, None, None

# ---------------- PLANUNG ----------------
if lehrer_file and klassenlehrer_file:
    # Lehrer-Klassen-Daten verarbeiten
    try:
        df_lehrer_raw = pd.read_csv(lehrer_file, sep=';', header=None)
        
        # Spaltennamen setzen basierend auf Anzahl der Spalten
        if len(df_lehrer_raw.columns) >= 6:
            df_lehrer_raw.columns = ["Spalte1", "Spalte2", "Spalte3", "Spalte4", "Klasse", "Lehrer"] + [f"Spalte{i}" for i in range(7, len(df_lehrer_raw.columns) + 1)]
        elif len(df_lehrer_raw.columns) == 2:
            df_lehrer_raw.columns = ["Klasse", "Lehrer"]
        else:
            st.error("Unerwartetes CSV-Format. Erwartet werden 2 oder mehr Spalten.")
            st.stop()
        
        # Nur relevante Spalten behalten und leere Zeilen entfernen
        df_lehrer = df_lehrer_raw[["Klasse", "Lehrer"]].dropna()
        
        # Leere Einträge entfernen und bereinigen
        df_lehrer = df_lehrer[df_lehrer["Klasse"].astype(str).str.strip() != ""]
        df_lehrer = df_lehrer[df_lehrer["Lehrer"].astype(str).str.strip() != ""]
        df_lehrer = df_lehrer[df_lehrer["Klasse"].astype(str) != "nan"]
        df_lehrer = df_lehrer[df_lehrer["Lehrer"].astype(str) != "nan"]
        
        # Klassen und Lehrer bereinigen
        df_lehrer["Klasse"] = df_lehrer["Klasse"].astype(str).str.strip()
        df_lehrer["Lehrer"] = df_lehrer["Lehrer"].astype(str).str.strip()

        # Klassenlehrer-Daten verarbeiten
        df_klassenlehrer = pd.read_csv(klassenlehrer_file)
        df_klassenlehrer.columns = df_klassenlehrer.columns.str.strip()
        klassenlehrer_cols = df_klassenlehrer.columns
        
        # Beide Klassenlehrer erfassen
        klassenlehrer_map = {}
        klassenlehrer_map_full = {}
        
        for _, row in df_klassenlehrer.iterrows():
            klasse = str(row[klassenlehrer_cols[0]]).strip()
            lehrer1 = str(row[klassenlehrer_cols[1]]).strip() if pd.notna(row[klassenlehrer_cols[1]]) else ""
            lehrer2 = str(row[klassenlehrer_cols[2]]).strip() if len(klassenlehrer_cols) > 2 and pd.notna(row[klassenlehrer_cols[2]]) else ""
            
            klassenlehrer_map[klasse] = lehrer1
            klassenlehrer_map_full[klasse] = [lehrer for lehrer in [lehrer1, lehrer2] if lehrer]
        
        st.session_state.klassenlehrer_map = klassenlehrer_map
        st.session_state.klassenlehrer_map_full = klassenlehrer_map_full
        
        # Parse Jahrgänge
        jahrgaenge = jahrgaenge_input.strip()
        if '-' in jahrgaenge:
            start, end = map(int, jahrgaenge.split('-'))
        else:
            start = end = int(jahrgaenge)
        sek_klassen = []
        for j in range(start, end + 1):
            for buchstabe in ['a', 'b', 'c', 'd', 'e']:
                klasse = f"{j}{buchstabe}"
                if klasse in df_lehrer["Klasse"].unique() or klasse in klassenlehrer_map_full:
                    sek_klassen.append(klasse)

        # Direkte Filterung basierend auf bereinigten Klassennamen
        def filter_klassen_direct(df, klassen_liste):
            df_clean = df.copy()
            return df[df_clean["Klasse"].isin(klassen_liste)]

        df_lehrer_1 = filter_klassen_direct(df_lehrer, sek_klassen)

        # Debug-Ausgabe zur Überprüfung (nur für Plan 1)
        klassen_1 = sorted(df_lehrer_1["Klasse"].unique())

        st.write(f"**Plan 1 (Jahrgänge {jahrgaenge_input}):** {len(klassen_1)} Klassen")
        if klassen_1:
            st.write(f"{', '.join(klassen_1)}")
        else:
            st.warning("Keine Klassen für Plan 1 gefunden!")

        # Parse Räume
        raeume_list = [r.strip() for r in raeume_input.split(",")]
        if len(raeume_list) != anzahl_parallel:
            st.error(f"Anzahl Räume ({len(raeume_list)}) stimmt nicht mit Parallelen ({anzahl_parallel}) überein.")
            st.stop()

        if st.button("🚀 Konferenzplan erzeugen"):
            # Nur Plan 1 erstellen
            with st.spinner("Erstelle Plan 1..."):
                (st.session_state.df_plan_1, st.session_state.df_warte_1, 
                 st.session_state.lehrer_einsaetze_1, st.session_state.df_klassensicht_1, 
                 st.session_state.statistik_1) = erstelle_konferenzplan(df_lehrer_1, klassenlehrer_map_full, 1, startzeit_input, dauer_konf, dauer_pause, anzahl_parallel, raeume_list)
            
            st.session_state.berechnung_abgeschlossen = True
            st.rerun()
            
    except Exception as e:
        st.error(f"Fehler beim Verarbeiten der Dateien: {str(e)}")
        st.write("Bitte überprüfen Sie das Format der CSV-Dateien.")

# ---------------- ANZEIGE UND EXPORT ----------------
if st.session_state.berechnung_abgeschlossen:
    st.success("✅ Konferenzplan erfolgreich erstellt!")
    
    if st.session_state.df_plan_1 is not None:
        # ZUSAMMENKLAPPBARE ABSCHNITTE
        with st.expander("👤 **Klassenlehrer-Zuordnung**", expanded=False):
            kl_data = []
            for klasse, lehrer_liste in sorted(st.session_state.klassenlehrer_map_full.items()):
                kl_data.append([klasse, ", ".join(lehrer_liste)])
            df_kl = pd.DataFrame(kl_data, columns=["Klasse", "Klassenlehrer"])
            st.dataframe(df_kl, width='stretch')
        
        with st.expander("📋 **Konferenzplan**", expanded=True):
            # Klassenlehrer in Lehrkräfte-Spalte fett machen (HTML)
            def format_lehrkraefte(row):
                teilnehmer = row["Lehrkräfte"].split(", ")
                klasse = row["Klasse"]
                kl_lehrer = st.session_state.klassenlehrer_map_full.get(klasse, [])
                formatted = []
                for l in teilnehmer:
                    if l.strip() in kl_lehrer:
                        formatted.append(f"<b>{l}</b>")
                    else:
                        formatted.append(l)
                return ", ".join(formatted)
            
            df_display = st.session_state.df_plan_1.copy()
            df_display["Lehrkräfte"] = df_display.apply(format_lehrkraefte, axis=1)
            st.markdown(df_display.to_html(escape=False, index=False), unsafe_allow_html=True)
        
        with st.expander("⏱️ **Wartezeiten je Lehrer**", expanded=False):
            st.dataframe(st.session_state.df_warte_1, width='stretch')
        
        with st.expander("👥 **Klassensicht** - Lehrer pro Klasse", expanded=False):
            st.dataframe(st.session_state.df_klassensicht_1, width='stretch')
        
        with st.expander("📊 **Statistiken**", expanded=False):
            for kategorie, daten in st.session_state.statistik_1.items():
                st.write(f"**{kategorie}**")
                stats_df = pd.DataFrame.from_dict(daten, orient="index", columns=["Wert"])
                st.dataframe(stats_df, width='stretch')
    else:
        st.warning("Kein Plan verfügbar")

    # Export-Optionen
    st.subheader("📤 Exportoptionen")
    
    styles = getSampleStyleSheet()
    bold_style = ParagraphStyle('BoldStyle', parent=styles['Normal'], fontName='Helvetica-Bold')

    # Funktion für Excel-Export (Standard)
    def export_plan_to_excel(df_plan, df_warte, df_klassensicht, statistik, plan_name):
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            df_plan.to_excel(writer, index=False, sheet_name="Konferenzplan")
            df_warte.to_excel(writer, index=False, sheet_name="Wartezeiten")
            
            # Klassensicht mit Lehrern als String
            dfk = df_klassensicht.copy()
            dfk["Lehrkräfte"] = dfk["Lehrkräfte"].apply(lambda x: ", ".join(x))
            dfk.to_excel(writer, index=False, sheet_name="Klassensicht")
            
            # Statistik in Excel exportieren
            for kategorie, daten in statistik.items():
                stats_df = pd.DataFrame.from_dict(daten, orient="index", columns=["Wert"])
                stats_df.to_excel(writer, sheet_name=f"Statistik_{kategorie[:10]}")
        
        return excel_buffer.getvalue()

    # Neue Funktion für Matrix-Excel-Export (wie PDF-Layout)
    def export_plan_to_matrix_excel(df_plan, lehrer_einsaetze, df_klassensicht, klassenlehrer_map_full, plan_name, raeume_list):
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            wb = writer.book

            # Zeit-Raum-Matrix
            ws_matrix = wb.create_sheet("Matrix")
            zeitslots_start = sorted(df_plan["Startzeit"].unique())
            zeitslots_end = sorted(df_plan["Endzeit"].unique())
            
            # Header
            ws_matrix.cell(1, 1, "Startzeit")
            ws_matrix.cell(1, 2, "Endzeit")
            for i, raum in enumerate(raeume_list, start=3):
                ws_matrix.cell(1, i, raum)
            
            for row_idx, (start, end) in enumerate(zip(zeitslots_start, zeitslots_end), start=2):
                ws_matrix.cell(row_idx, 1, start)
                ws_matrix.cell(row_idx, 2, end)
                df_slot = df_plan[df_plan["Startzeit"] == start]
                for _, rowdata in df_slot.iterrows():
                    raum_idx = raeume_list.index(rowdata["Raum"]) + 3
                    klasse = rowdata["Klasse"]
                    kl_lehrer = klassenlehrer_map_full.get(klasse, [])
                    kl_str = f" ({', '.join(kl_lehrer)})" if kl_lehrer else ""
                    ws_matrix.cell(row_idx, raum_idx, f"{klasse}{kl_str}")
            
            # Lehrereinsatzübersicht
            ws_lehrer = wb.create_sheet("Lehrereinsatz")
            ws_lehrer.page_setup.orientation = 'landscape'
            ws_lehrer.cell(1, 1, "Lehrer")
            for col_idx, start in enumerate(zeitslots_start, start=2):
                ws_lehrer.cell(1, col_idx, start)
            
            for row_idx, (l, einsaetze) in enumerate(lehrer_einsaetze.items(), start=2):
                ws_lehrer.cell(row_idx, 1, l)
                for z, k, r in einsaetze:
                    col_idx = zeitslots_start.index(z) + 2
                    is_klassenlehrer = k in klassenlehrer_map_full and l in klassenlehrer_map_full[k]
                    cell_value = f"{k} ({r})"
                    ws_lehrer.cell(row_idx, col_idx, cell_value)
                    if is_klassenlehrer:
                        cell = ws_lehrer.cell(row_idx, col_idx)
                        cell.font = openpyxl.styles.Font(bold=True)
            
            # Klassensicht
            ws_klasse = wb.create_sheet("Klassensicht")
            ws_klasse.page_setup.orientation = 'landscape'
            ws_klasse.cell(1, 1, "Klasse")
            ws_klasse.cell(1, 2, "Lehrkräfte")
            for row_idx, (_, row) in enumerate(df_klassensicht.iterrows(), start=2):
                ws_klasse.cell(row_idx, 1, row["Klasse"])
                lehrer_liste = []
                for lehrer in row["Lehrkräfte"]:
                    is_klassenlehrer = row["Klasse"] in klassenlehrer_map_full and lehrer in klassenlehrer_map_full[row["Klasse"]]
                    if is_klassenlehrer:
                        lehrer_liste.append(f"{lehrer} (KL)")
                    else:
                        lehrer_liste.append(lehrer)
                ws_klasse.cell(row_idx, 2, ", ".join(lehrer_liste))
            
            # Umrandungen für alle Sheets hinzufügen
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for ws in wb:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            cell.border = thin_border
        
        return excel_buffer.getvalue()

    # Funktion für PDF-Export
    def export_plan_to_pdf(df_plan, lehrer_einsaetze, df_klassensicht, klassenlehrer_map_full, plan_name, raeume_list):
        pdf_buffer = BytesIO()
        doc = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4),
                                leftMargin=20, rightMargin=20,
                                topMargin=30, bottomMargin=30)
        elements = []

        elements.append(Paragraph(f"Konferenzplan - {plan_name}", styles["Title"]))
        elements.append(Spacer(1, 12))

        # Zeit-Raum-Matrix
        zeitslots_start = sorted(df_plan["Startzeit"].unique())
        zeitslots_end = sorted(df_plan["Endzeit"].unique())
        
        # Spaltenbreiten anpassen für besseres Layout
        page_width, page_height = landscape(A4)
        available_width = page_width - 40
        col_widths = [60, 60]
        for i in range(len(raeume_list)):
            col_widths.append((available_width - 120) / len(raeume_list))
        
        matrix_data = [
            [Paragraph("Startzeit", styles["Normal"]), 
             Paragraph("Endzeit", styles["Normal"])] + 
            [Paragraph(raum, styles["Normal"]) for raum in raeume_list]
        ]
        
        for start, end in zip(zeitslots_start, zeitslots_end):
            row = [Paragraph(start, styles["Normal"]), Paragraph(end, styles["Normal"])] + [""] * len(raeume_list)
            df_slot = df_plan[df_plan["Startzeit"] == start]
            
            for _, rowdata in df_slot.iterrows():
                r = raeume_list.index(rowdata["Raum"])
                klasse = rowdata["Klasse"]
                kl_lehrer = klassenlehrer_map_full.get(klasse, [])
                kl_str = f" ({', '.join(kl_lehrer)})" if kl_lehrer else ""
                # Prüfen ob Klassenlehrer (Haupt oder Stellvertreter)
                is_klassenlehrer = klasse in klassenlehrer_map_full and any(
                    kl_lehrer in rowdata["Lehrkräfte"] 
                    for kl_lehrer in klassenlehrer_map_full[klasse]
                )
                cell_text = f"{klasse}{kl_str}"
                if is_klassenlehrer:
                    row[r + 2] = Paragraph(f"<b>{cell_text}</b>", styles["Normal"])
                else:
                    row[r + 2] = Paragraph(cell_text, styles["Normal"])
            
            matrix_data.append(row)
        
        table = Table(matrix_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))
        elements.append(table)
        elements.append(PageBreak())

        # Lehrereinsatzübersicht
        elements.append(Paragraph("Lehrereinsatzübersicht", styles["Title"]))
        elements.append(Spacer(1, 12))
        
        available_width_teacher = page_width - 40
        lehrer_col_widths = [70]
        slot_width = (available_width_teacher - 70) / len(zeitslots_start)
        for _ in zeitslots_start:
            lehrer_col_widths.append(min(slot_width, 100))
        
        header_row = [Paragraph("Lehrer", styles["Normal"])]
        for slot in zeitslots_start:
            header_row.append(Paragraph(slot, styles["Normal"]))
        lehrertabelle = [header_row]
        
        for l, einsaetze in lehrer_einsaetze.items():
            reihe = [Paragraph(l, styles["Normal"])]
            
            for slot in zeitslots_start:
                klassen_in_slot = []
                for z, k, r in einsaetze:
                    if z == slot:
                        # Prüfen ob Klassenlehrer
                        is_klassenlehrer = k in klassenlehrer_map_full and l in klassenlehrer_map_full[k]
                        if is_klassenlehrer:
                            klassen_in_slot.append(f"<b>{k}</b>")
                        else:
                            klassen_in_slot.append(k)
                
                if klassen_in_slot:
                    klassen_text = ", ".join(klassen_in_slot)
                    reihe.append(Paragraph(klassen_text, styles["Normal"]))
                else:
                    reihe.append(Paragraph("", styles["Normal"]))
            
            lehrertabelle.append(reihe)
        
        t2 = Table(lehrertabelle, colWidths=lehrer_col_widths, repeatRows=1)
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.beige),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))
        elements.append(t2)
        elements.append(PageBreak())

        # Klassensicht
        elements.append(Paragraph("Klassensicht - Lehrer pro Klasse", styles["Title"]))
        elements.append(Spacer(1, 12))
        
        data_klassensicht = [["Klasse", "Lehrkräfte"]]
        for _, row in df_klassensicht.iterrows():
            lehrer_liste = []
            for lehrer in row["Lehrkräfte"]:
                # Prüfen ob Klassenlehrer
                is_klassenlehrer = row["Klasse"] in klassenlehrer_map_full and lehrer in klassenlehrer_map_full[row["Klasse"]]
                if is_klassenlehrer:
                    lehrer_liste.append(f"<b>{lehrer}</b>")
                else:
                    lehrer_liste.append(lehrer)
            
            klasse_cell = Paragraph(row["Klasse"], styles["Normal"])
            lehrer_cell = Paragraph(", ".join(lehrer_liste), styles["Normal"])
            data_klassensicht.append([klasse_cell, lehrer_cell])
        
        table_klassensicht = Table(data_klassensicht, repeatRows=1, colWidths=[100, available_width - 100])
        table_klassensicht.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("WORDWRAP", (0, 0), (-1, -1), True),
        ]))
        elements.append(table_klassensicht)
        
        doc.build(elements)
        return pdf_buffer.getvalue()

    # Export-Buttons
    if st.session_state.df_plan_1 is not None:
        
        # Excel Export (Standard)
        excel_data_1 = export_plan_to_excel(
            st.session_state.df_plan_1, 
            st.session_state.df_warte_1,
            st.session_state.df_klassensicht_1,
            st.session_state.statistik_1,
            "Konferenzplan"
        )
        st.download_button("📥 Excel (Daten)", excel_data_1, "konferenzplan_daten.xlsx", use_container_width=True)
        
        # Matrix Excel Export
        matrix_excel_data_1 = export_plan_to_matrix_excel(
            st.session_state.df_plan_1, 
            st.session_state.lehrer_einsaetze_1,
            st.session_state.df_klassensicht_1,
            st.session_state.klassenlehrer_map_full,
            "Konferenzplan",
            raeume_list
        )
        st.download_button("📥 Excel (Layout)", matrix_excel_data_1, "konferenzplan_layout.xlsx", use_container_width=True)
        
        # PDF Export
        if st.button("📄 PDF erstellen", use_container_width=True, key="pdf1"):
            with st.spinner("PDF wird erstellt..."):
                pdf_data_1 = export_plan_to_pdf(
                    st.session_state.df_plan_1, 
                    st.session_state.lehrer_einsaetze_1,
                    st.session_state.df_klassensicht_1,
                    st.session_state.klassenlehrer_map_full,
                    "Konferenzplan",
                    raeume_list
                )
                st.download_button("📄 PDF herunterladen", pdf_data_1, "konferenzplan.pdf", use_container_width=True)
